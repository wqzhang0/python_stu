# from xlrd import open_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import PatternFill

secondary_excel_path = "D:\PYTHON\\2019年定做登记\\1月份定做\\1月1号定做空调门、拉链、整扇、.xlsx"

from openpyxl import load_workbook, Workbook


class OrderInfo:
    def __init__(self, x, y, code, catalog):
        self.x = x
        self.y = y
        self.code = code
        self.catalog = catalog
        self.preform = None

    def set_preform(self, msg):
        self.preform = msg

    def __str__(self):
        if self.preform:
            return f"编号：{self.code} 宽：{self.x} 高:{self.y} 分类：{self.catalog} 辅助信息：{self.preform}"
        return f"编号：{self.code} 宽：{self.x} 高:{self.y} 分类：{self.catalog}"


# wb = Workbook(secondary_excel_path)
wb = load_workbook(secondary_excel_path)
print(wb.sheetnames)

ws2 = wb.get_sheet_by_name("Sheet2")
wb_r = Workbook()
if not "Sheet_R" in wb_r.sheetnames:
    ws_r = wb_r.create_sheet("Sheet_R")
else:
    ws_r = wb_r.get_sheet_by_name("Sheet_R")
print(ws2.max_column)
print(ws2.max_row)
current_catalog_name = ""

# 判断总函数,按照5行 递增
max_column = ws2.max_column
max_row = ws2.max_row
index_column = 1
order_list = []

code_col_index_set = set()
for _row in range(max_row):
    for _col in range(max_column):
        c_cell = ws2.cell(row=_row + 1, column=_col + 1).value
        c_cell2 = ws2.cell(row=_row + 1, column=_col + 2).value
        c_cell3 = ws2.cell(row=_row + 1, column=_col + 3).value
        if c_cell and c_cell2 and c_cell3:
            if isinstance(c_cell, str) and isinstance(c_cell2, str) and isinstance(c_cell3, str):

                if c_cell.strip() == '编号' and c_cell2.strip() == '宽度' and c_cell3.strip() == '高度':
                    print(ws2.cell(row=_row + 1, column=_col + 1))
                    code_col_index_set.add(_col + 1)
# fill = PatternFill("solid", fgColor="1874CD")
fill = PatternFill("solid", fgColor="1874CD")
# FILL_SOLID
for code_col_index in code_col_index_set:
    for cells in ws2.iter_cols(min_row=0, min_col=code_col_index, max_col=code_col_index + 3, max_row=max_row):
        all_blank = True
        for cell in cells:
            if cell.value:
                all_blank = False
        if not all_blank:
            for cell in cells:
                cell.fill = fill

secondary_excel_path2 = "D:\PYTHON\\2019年定做登记\\1月份定做\\1月1号定做空调门、拉链、整扇、oot.xlsx"

wb.save(secondary_excel_path2)
# for row in ws2.iter_rows(min_row=0, min_col=0, max_col=max_column, max_row=max_row):
#     # is_pre_preffix = False
#     row.
#     for i, cell in enumerate(row):
#         if
#         if cell.value == ' 编号	宽度	高度'
#         print(cell)
# while index_column < max_column:
#
#     # 判断哪些是有效行数
#     #
#
#
#
#     for row in ws2.iter_rows(min_row=0, min_col=index_column, max_col=index_column + 5, max_row=max_row):
#         # 方块
#         is_catalog_cell = False
#         merge_count = 0
#         is_pre_preffix = False
#         for cell in row:
#             if isinstance(cell, MergedCell):
#                 merge_count += 1
#         if merge_count == 2:
#             is_catalog_cell = True
#         if merge_count == 3:
#             is_pre_preffix = True
#         if is_catalog_cell:
#             current_catalog_name = row[0].value
#             continue
#         elif is_pre_preffix:
#             pre_order = order_list[-1]
#             pre_order.set_preform(row[0].value)
#             continue
#         else:
#
#             if isinstance(row[0].value, str) and row[0].value.strip() == '编号':
#                 continue
#             if not row[0].value:
#                 continue
#             order = OrderInfo(x=row[1].value, y=row[2].value, code=row[0].value, catalog=current_catalog_name)
#             order_list.append(order)
#     index_column += 5

# d2 = ws2.cell(row=4, column=2)
# print(d2)
# d1 = ws2.cell(row=4, column=5, value=10)
# print(d1)
#
# print(ws2['A4'])
# cell_range = ws2['A1':'C2']


# wb.save('smp.xlsx')
# grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# ws2 = wb.create_sheet('oh ny zsh',2)
# ws2.title = 'Nwe Ttitle'
# ws2.sheet_properties.tabColor = "1072Ba"
#
# # wb.save("sample2.xlsx")


#
