# from xlrd import open_workbook
from openpyxl.cell import Cell, MergedCell

secondary_excel_path = "G:\SynologyDrive\文档\人大项目\各种业务导入导出模板\会议信息.xlsx"

from openpyxl import load_workbook


class OrderInfo:
    def __init__(self, x, y, code, catelog):
        self.x = x
        self.y = y
        self.code = code
        self.catelog = catelog


# wb = Workbook(secondary_excel_path)
wb = load_workbook(secondary_excel_path)
print(wb.sheetnames)

ws2 = wb.get_sheet_by_name("Sheet1")

if not "Sheet_R" in wb.sheetnames:
    wb_r = wb.create_sheet("Sheet_R")
else:
    wb_r = wb.get_sheet_by_name("Sheet_R")
print('------------------------------')
print(ws2.max_column)
print(ws2.max_row)
catalog_name = ""

# 判断总函数,按照5行 递增
max_column  = ws2.max_column
max_row  = ws2.max_row
index_column = 1
for row in ws2.iter_rows(min_row=index_column,max_col=5,max_row=max_row):
    # 方块
    pass
for row in ws2.iter_rows(min_row=1, max_col=3, max_row=200):
    for cell in row:
        if isinstance(cell, Cell):
            print("非合并")
        elif isinstance(cell, MergedCell):
            print("合并")
        else:
            raise Exception("sss")
        print(cell)
# d2 = ws2.cell(row=4, column=2)
# print(d2)
# d1 = ws2.cell(row=4, column=5, value=10)
# print(d1)
#
# print(ws2['A4'])
# cell_range = ws2['A1':'C2']


# wb.save('smp.xlsx')
# grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# ws2 = wb.create_sheet('oh ny zsh',2)
# ws2.title = 'Nwe Ttitle'
# ws2.sheet_properties.tabColor = "1072Ba"
#
# # wb.save("sample2.xlsx")
