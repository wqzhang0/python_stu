# from xlrd import open_workbook


from openpyxl import load_workbook, Workbook


def reade_excel(secondary_excel_path, all_catalogs):
    wb = load_workbook(secondary_excel_path)
    print(wb.sheetnames)
    ws2 = wb.get_sheet_by_name("Sheet2")

    print('------------------------------')
    print(ws2.max_column)
    print(ws2.max_row)
    current_catalog_name = ""

    # 判断总函数,按照5行 递增
    max_column = ws2.max_column
    max_row = ws2.max_row

    for row in ws2.iter_rows(min_row=0, min_col=0, max_col=max_column, max_row=max_row):
        all_catalogs.add(row[0].value)


if __name__ == '__main__':
    # import pandas as pd
    # import numpy as np
    #
    # data = {
    #     'apples': [3, 2, 0, 1],
    #     'oranges': [0, 3, 7, 2]
    # }
    # purchases = pd.DataFrame(data,index=['June', 'Robert', 'Lily', 'David'])
    #
    # print(purchases)
    # print(purchases.loc['June'])
    from pathlib import Path, PurePath

    # base_path = PurePath("F:")
    base_path = Path("D:\PYTHON\\2019年定做登记")
    print(base_path.absolute())
    print(base_path.cwd())

    all_catalogs = set()
    file_num = 0

    for time_dir in base_path.iterdir():
        if time_dir.is_dir():
            # for file in time_dir.iterdir() :
            for file in time_dir.glob('*.xlsx'):
                print(file)
                file_num +=1
                reade_excel(file, all_catalogs)
    wb_r = Workbook()

    if not "Sheet_R" in wb_r.sheetnames:
        ws_r = wb_r.create_sheet("Sheet_R")
    else:
        ws_r = wb_r.get_sheet_by_name("Sheet_R")


    for i, o in enumerate(all_catalogs.difference(set(range(0,10000)))):
        ws_r.cell(column=1, row=i + 1, value=o)
    print(f"读取文件数量：{file_num}")

    wb_r.save('catalog.xlsx')
