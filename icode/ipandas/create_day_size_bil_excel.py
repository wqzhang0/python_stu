# from xlrd import open_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
    BubbleChart)
from openpyxl.chart.layout import Layout, ManualLayout

from openpyxl import load_workbook, Workbook


class OrderInfo:
    def __init__(self, x, y, code, catalog):
        self.x = x
        self.y = y
        self.code = code
        self.catalog = catalog
        self.preform = None
        self.file_name = ""

    def set_preform(self, msg):
        self.preform = msg

    def set_file_name(self, file_name):
        self.file_name = file_name

    def __str__(self):
        if self.preform:
            return f"编号：{self.code} 宽：{self.x} 高:{self.y} 分类：{self.catalog} 辅助信息：{self.preform}"
        return f"编号：{self.code} 宽：{self.x} 高:{self.y} 分类：{self.catalog}"


wb_r = Workbook()
ws_r = wb_r.active
order_list = []


def create_chart_excel(secondary_excel_path):
    # wb = Workbook(secondary_excel_path)
    wb = load_workbook(secondary_excel_path)
    print(wb.sheetnames)
    ws2 = wb.get_sheet_by_name("Sheet2")
    current_catalog_name = ""
    # 判断总函数,按照5行 递增
    max_column = ws2.max_column
    max_row = ws2.max_row
    index_column = 1
    while index_column < max_column:
        for row in ws2.iter_rows(min_row=0, min_col=index_column, max_col=index_column + 5, max_row=max_row):
            # 方块
            is_catalog_cell = False
            merge_count = 0
            is_pre_preffix = False
            for cell in row:
                if isinstance(cell, MergedCell):
                    merge_count += 1
            if merge_count == 2:
                is_catalog_cell = True
            if merge_count == 3:
                is_pre_preffix = True
            if is_catalog_cell:
                current_catalog_name = row[0].value
                continue
            elif is_pre_preffix:
                pre_order = order_list[-1]
                pre_order.set_preform(row[0].value)
                continue
            else:

                if isinstance(row[0].value, str) and row[0].value.strip() == '编号':
                    continue
                if not row[0].value:
                    continue
                order = OrderInfo(x=row[1].value, y=row[2].value, code=row[0].value, catalog=current_catalog_name)

                order.set_file_name(secondary_excel_path.name)
                order_list.append(order)

        index_column += 5


if __name__ == '__main__':
    from pathlib import Path, PurePath

    # base_path = PurePath("F:")
    base_path = Path("D:\PYTHON\\2019年定做登记")

    print(base_path.absolute())
    print(base_path.cwd())

    all_catalogs = set()
    file_num = 0

    for time_dir in base_path.iterdir():
        if time_dir.is_dir():
            # for file in time_dir.iterdir() :
            for file in time_dir.glob('*.xlsx'):
                print(file)
                file_num += 1
                create_chart_excel(file)

    hight_list = [o.y if o and isinstance(o.y,int) else 0  for o in order_list]
    low_num = min(hight_list)
    hight_dict = {}
    while low_num <= max(hight_list):
        for hight in hight_list:
            if hight > low_num and hight <= low_num + 5:
                hight_dict[f'{low_num}-{low_num + 5}'] = hight_dict.get(f'{low_num}-{low_num + 5}', 0)

    print(hight_dict)

    row_index = 1
    for k, v in hight_dict.items():
        ws_r.cell(column=1, row=row_index, value=k)
        ws_r.cell(column=2, row=row_index, value=v)
        row_index += 1
    wb_r.save("create_day_size_bili_excel.xlsx")
